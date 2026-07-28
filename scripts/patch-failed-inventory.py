"""
SDOT File Inventory Patcher (patch-failed-inventory.py)

목적: 이미 생성된 엑셀 인벤토리에서 '성공'하지 못한 파일(8개)만 찾아내어,
파일명 정규식을 통해 start_date와 end_date를 강제로 채워 넣고 엑셀을 업데이트합니다.
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import re

# 프로젝트 최상위 디렉토리 및 엑셀 파일 경로 동적 계산
PROJECT_ROOT = Path(__file__).resolve().parent.parent
excel_path = PROJECT_ROOT / 'metadata' / 'sdot-schema-mapping.xlsx'

def extract_dates_from_filename(filename: str):
    """파일명에서 2021.01.25-01.31 형태를 추출하여 시작/종료일로 변환"""
    pattern = r'(\d{4})\.(\d{2})\.(\d{2})-(\d{2})\.(\d{2})'
    match = re.search(pattern, filename)
    
    if match:
        year, start_month, start_day, end_month, end_day = match.groups()
        start_date = f"{year}-{start_month}-{start_day} 00:00:00"
        
        end_year = year
        if int(start_month) == 12 and int(end_month) == 1:
            end_year = str(int(year) + 1)
            
        end_date = f"{end_year}-{end_month}-{end_day} 23:59:59"
        return start_date, end_date
    return None, None

def calculate_period_type(s_date: str, e_date: str) -> str:
    """날짜 문자열을 바탕으로 데이터의 기간 단위를 판별"""
    try:
        start = pd.to_datetime(s_date)
        end = pd.to_datetime(e_date)
        days = (end - start).days
        
        if days >= 300: return 'yearly'
        elif days >= 25: return 'monthly'
        elif days >= 6: return 'weekly'
        else: return 'daily'
    except Exception:
        return 'unknown'

def patch_excel():
    if not excel_path.exists():
        print(f"🚨 에러: 엑셀 파일을 찾을 수 없습니다! ({excel_path})")
        return
        
    print(f"📂 엑셀 파일 로드 중... ({excel_path.name})")
    
    # 1. 기존 데이터 읽기
    df = pd.read_excel(excel_path, sheet_name='FileInventory')
    
    # 2. 실패한 행 찾기 [수정됨: '성공'과 정확히 일치하지 않는 모든 행 선택]
    failed_mask = df['remarks'] != '성공'
    failed_count = failed_mask.sum()
    
    print(f"🔍 복구 대상 파일 {failed_count}개를 찾았습니다.\n")
    
    if failed_count == 0:
        print("✅ 이미 모든 파일이 성공적으로 파싱되어 있습니다.")
        return

    # 3. 실패한 행에 대해서만 파일명에서 날짜 추출 및 업데이트
    for idx, row in df[failed_mask].iterrows():
        fname = row['file_name']
        target_col = row['target_time_col']
        s_date, e_date = extract_dates_from_filename(fname)
        
        if s_date and e_date:
            df.at[idx, 'start_date'] = s_date
            df.at[idx, 'end_date'] = e_date
            df.at[idx, 'period_type'] = calculate_period_type(s_date, e_date)
            df.at[idx, 'remarks'] = f"성공 (데이터 훼손으로 파일명에서 복구)"
            print(f"  ✅ [복구 완료] {fname}\n     -> {s_date} ~ {e_date}")
        else:
            print(f"  ❌ [복구 실패] {fname}: 정규식 패턴 불일치")

    # 4. 엑셀 파일 덮어쓰기 (FileInventory 시트만 교체하고 나머지 시트는 보존)
    print("\n💾 엑셀 파일 시트 업데이트 중...")
    wb = openpyxl.load_workbook(excel_path)
    
    if 'FileInventory' in wb.sheetnames:
        del wb['FileInventory']
    ws = wb.create_sheet('FileInventory', index=0)
    
    # DataFrame을 Excel 시트에 쓰기
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    # 헤더 볼드체 처리
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(excel_path)
    print(f"🎉 완벽합니다! '{excel_path.name}'의 부분 패치가 완료되었습니다.")

if __name__ == "__main__":
    patch_excel()