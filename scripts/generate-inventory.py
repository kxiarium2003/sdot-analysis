"""
SDOT File Inventory Generator (generate-inventory.py) - V3

주요 개선 사항
1. I/O 최적화: 메타데이터 및 날짜(Date) 컬럼만 부분 로드하여 메모리 사용량 최소화
2. 도메인 로직: S-DoT 연도별 스키마 버전 명시적 분류 (측정시간 > 등록일자 > 전송시간 최우선 탐색)
3. 주기 판별: 데이터의 총 기간 기반으로 daily/monthly/yearly 판별
4. 날짜 파싱 강화: 마침표(.), 언더바(_), 하이픈(-) 등 다중 포맷 완벽 호환 및 결측치 방어
5. 데이터 무결성: index_col=False 옵션으로 CSV 콤마 오작동(데이터 밀림 현상) 방어
6. 메타데이터 보존: 파일 내 존재하는 모든 원본 컬럼(all_columns_list) 박제
7. 경로 강건성: 스크립트 실행 위치와 무관하게 프로젝트 절대 경로 자동 추적
"""

import argparse
import logging
import hashlib
import sys
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# 지원하는 인코딩 목록
SUPPORTED_ENCODINGS = ['utf-8', 'cp949', 'euc-kr', 'utf-8-sig']

# 프로젝트 최상위 디렉토리 절대 경로 동적 계산 (scripts 폴더의 부모 폴더)
PROJECT_ROOT = Path(__file__).resolve().parent.parent

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="SDOT CSV 파일들의 최적화된 인벤토리를 생성합니다.")
    parser.add_argument(
        '--raw-dir', 
        type=str, 
        default=str(PROJECT_ROOT / 'data' / 'raw'),
        help='Raw CSV 파일들이 위치한 디렉토리 경로'
    )
    parser.add_argument(
        '--output', 
        type=str, 
        default=str(PROJECT_ROOT / 'metadata' / 'sdot-schema-mapping.xlsx'),
        help='메타데이터를 저장할 엑셀 파일 경로'
    )
    return parser.parse_args()

def setup_logging() -> logging.Logger:
    logger = logging.getLogger('generate-inventory')
    logger.setLevel(logging.INFO)
    
    if not logger.handlers:
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        ch = logging.StreamHandler(sys.stdout)
        ch.setFormatter(formatter)
        logger.addHandler(ch)
        
    return logger

def extract_year(filename: str) -> Optional[int]:
    """파일명에서 202x 형태의 연도를 추출합니다."""
    match = re.search(r'(202\d{1})', filename)
    return int(match.group(1)) if match else None

def inspect_csv_header(file_path: Path, logger: logging.Logger) -> Tuple[str, List[str]]:
    """
    pandas 파서를 이용하여 안전하게 인코딩과 컬럼 목록을 추출합니다.
    [핵심 보완] index_col=False를 추가하여 데이터 밀림(Shift) 현상을 원천 차단합니다.
    """
    for enc in SUPPORTED_ENCODINGS:
        try:
            df_head = pd.read_csv(file_path, encoding=enc, nrows=100, dtype=str, index_col=False)
            return enc, list(df_head.columns)
        except (UnicodeDecodeError, pd.errors.ParserError):
            continue
            
    logger.error(f"[{file_path.name}] 호환되는 인코딩을 찾을 수 없거나 파싱 에러가 발생했습니다.")
    raise ValueError(f"Failed to read CSV header: {file_path}")

def classify_schema(columns: List[str]) -> Tuple[str, Optional[str], str]:
    """
    [핵심 보완] 컬럼 목록을 바탕으로 시간 기준 컬럼을 최우선순위로 탐색합니다.
    우선순위: 1. 측정시간 -> 2. 등록일자 -> 3. 전송시간
    """
    col_str = ",".join(columns)
    column_hash = hashlib.md5(col_str.encode('utf-8')).hexdigest()[:8]
    
    date_col = None
    if '측정시간' in columns:
        schema_version = 'v2023_onward'
        date_col = '측정시간'
    elif '등록일자' in columns:
        schema_version = 'v2020_2022_recovered'
        date_col = '등록일자'
    elif '전송시간' in columns:
        schema_version = 'v2020_2022_corrupted'
        date_col = '전송시간'
    else:
        schema_version = f'v_unknown_{column_hash}'
        
    return schema_version, date_col, column_hash

def calculate_period_type(start_date: pd.Timestamp, end_date: pd.Timestamp) -> str:
    """총 기간(Date Diff)을 바탕으로 데이터의 기간 단위를 판별합니다."""
    if pd.isna(start_date) or pd.isna(end_date):
        return 'unknown'
        
    days = (end_date - start_date).days
    
    if days >= 300:
        return 'yearly'
    elif days >= 25:
        return 'monthly'
    elif days >= 6:
        return 'weekly'
    else:
        return 'daily'

def process_file(file_path: Path, raw_dir: Path, logger: logging.Logger) -> Optional[Dict[str, Any]]:
    """단일 파일에 대한 메타데이터 추출 작업을 수행합니다."""
    try:
        # 1. 헤더만 읽어서 인코딩, 컬럼 리스트(모든 컬럼), 스키마 분석
        encoding, columns = inspect_csv_header(file_path, logger)
        schema_version, date_col, column_hash = classify_schema(columns)
        
        row_count = 0
        start_date_str = None
        end_date_str = None
        period_type = 'unknown'
        remarks = '성공'
        
        # 2. 선택된 시간 컬럼만 로드하여 메모리 최적화 (데이터 밀림 방지를 위해 index_col=False 필수)
        if date_col:
            df_dates = pd.read_csv(file_path, encoding=encoding, usecols=[date_col], dtype=str, index_col=False)
            row_count = len(df_dates)
            
            if row_count > 0:
                # [핵심 보완] 유니버설 날짜 파싱 로직 적용
                raw_dates = df_dates[date_col].dropna().astype(str).str.strip()
                
                # '2025-03-18_00:07:00' 형태 처리 (언더바 제거)
                raw_dates = raw_dates.str.replace('_', ' ', regex=False)
                # 엑셀 지수 변환 찌꺼기 처리 ('202000000000.0' -> '202000000000')
                raw_dates = raw_dates.str.replace(r'\.0$', '', regex=True)
                
                # 1차 파싱: 하이픈(-), 마침표(.), 콜론(:)이 포함된 정상/혼합 포맷 처리
                parsed_dates = pd.to_datetime(raw_dates, errors='coerce')
                
                # 2차/3차 파싱: 순수 숫자로만 이루어진 S-DoT 포맷 구제 (12자리/14자리)
                if parsed_dates.isna().mean() > 0.5:
                    parsed_dates = pd.to_datetime(raw_dates, format='%Y%m%d%H%M', errors='coerce')
                if parsed_dates.isna().mean() > 0.5:
                    parsed_dates = pd.to_datetime(raw_dates, format='%Y%m%d%H%M%S', errors='coerce')

                valid_dates = parsed_dates.dropna()
                
                if not valid_dates.empty:
                    s_date = valid_dates.min()
                    e_date = valid_dates.max()
                    start_date_str = s_date.strftime('%Y-%m-%d %H:%M:%S')
                    end_date_str = e_date.strftime('%Y-%m-%d %H:%M:%S')
                    period_type = calculate_period_type(s_date, e_date)
                else:
                    remarks = f"날짜 데이터 파싱 불가 ({date_col} 포맷 훼손)"
        else:
            # Date 컬럼이 없는 알 수 없는 스키마의 경우 줄 수만 계산
            with open(file_path, 'r', encoding=encoding) as f:
                row_count = sum(1 for _ in f) - 1 # 헤더 제외
            remarks = "시간 관련 컬럼 존재하지 않음"
            
        return {
            'year': extract_year(file_path.name),
            'relative_path': str(file_path.relative_to(raw_dir)),
            'file_name': file_path.name,
            'file_size_mb': round(file_path.stat().st_size / (1024 * 1024), 2),
            'encoding': encoding,
            'row_count': row_count,
            'col_count': len(columns),
            'all_columns_list': ", ".join(columns),  # [추가] 파일 내 모든 컬럼 보존
            'column_hash': column_hash,
            'schema_version': schema_version,
            'target_time_col': date_col,             # [추가] 실제 파싱에 사용된 타겟 컬럼명
            'period_type': period_type,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'remarks': remarks
        }
        
    except Exception as e:
        logger.error(f"[{file_path.name}] 파일 처리 중 오류 발생: {str(e)}")
        return None

def build_inventory(raw_dir: Path, logger: logging.Logger) -> pd.DataFrame:
    """전체 디렉토리를 순회하며 인벤토리 데이터프레임을 구축합니다."""
    records = []
    csv_files = list(raw_dir.rglob('*.csv'))
    
    if not csv_files:
        logger.warning(f"[{raw_dir}] 경로에 CSV 파일이 존재하지 않습니다.")
        return pd.DataFrame()
        
    for idx, file_path in enumerate(csv_files, 1):
        logger.info(f"Processing ({idx}/{len(csv_files)}): {file_path.name}")
        record = process_file(file_path, raw_dir, logger)
        if record:
            records.append(record)
            
    df = pd.DataFrame(records)
    
    if not df.empty:
        df = df.sort_values(by=['year', 'file_name']).reset_index(drop=True)
        df['file_order'] = df.groupby('year').cumcount() + 1
        
        # [수정] 메타데이터 보존을 위해 새로 추가한 컬럼들을 순서에 맞게 배치
        cols_order = [
            'year', 'file_order', 'relative_path', 'file_name', 'file_size_mb',
            'encoding', 'row_count', 'col_count', 'all_columns_list', 'column_hash', 
            'schema_version', 'target_time_col', 'period_type', 'start_date', 'end_date', 'remarks'
        ]
        df = df[cols_order]
        
    return df

def update_excel(df_inventory: pd.DataFrame, output_path: Path, logger: logging.Logger) -> None:
    """기존 엑셀 파일의 다른 시트는 보존하고 'FileInventory' 시트만 덮어씁니다."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = 'FileInventory'
    
    if output_path.exists():
        logger.info(f"기존 엑셀 파일을 로드합니다: {output_path}")
        wb = openpyxl.load_workbook(output_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, index=0)
    else:
        logger.info("새로운 메타데이터 엑셀 파일을 생성합니다.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

    for r in dataframe_to_rows(df_inventory, index=False, header=True):
        ws.append(r)
        
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    wb.save(output_path)
    logger.info(f"✅ 엑셀 파일 업데이트 완료: {output_path} (시트명: {sheet_name})")

def print_summary(df_inventory: pd.DataFrame, logger: logging.Logger) -> None:
    if df_inventory.empty:
        logger.info("생성된 인벤토리가 없습니다.")
        return
        
    total_files = len(df_inventory)
    total_rows = df_inventory['row_count'].sum()
    schema_counts = df_inventory['schema_version'].value_counts().to_dict()
    failed_counts = len(df_inventory[df_inventory['remarks'] != '성공'])
    
    print("\n" + "="*50)
    print(" SDOT File Inventory Generation Summary")
    print("="*50)
    print(f"Total processed files : {total_files} files")
    print(f"Total rows (records)  : {total_rows:,} rows")
    print(f"Failed parsing files  : {failed_counts} files")
    print("-" * 50)
    print("Files per Schema Version:")
    for schema, count in schema_counts.items():
        print(f"  - {schema}: {count} files")
    print("="*50 + "\n")

def main() -> None:
    args = parse_args()
    logger = setup_logging()
    
    raw_dir = Path(args.raw_dir)
    output_path = Path(args.output)
    
    if not raw_dir.exists() or not raw_dir.is_dir():
        logger.error(f"원본 디렉토리를 찾을 수 없습니다: {raw_dir.resolve()}")
        sys.exit(1)
        
    logger.info(f"SDOT File Inventory 생성을 시작합니다. (Target: {raw_dir})")
    
    df_inventory = build_inventory(raw_dir, logger)
    
    if not df_inventory.empty:
        update_excel(df_inventory, output_path, logger)
        
    print_summary(df_inventory, logger)
    
if __name__ == "__main__":
    main()